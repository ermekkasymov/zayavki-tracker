# -*- coding: utf-8 -*-
"""Генератор файла 'Трекер_заявок.xlsx' по PRD.md"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, Reference

N_ROWS = 100
FIRST_ROW = 2
LAST_ROW = FIRST_ROW + N_ROWS - 1  # 101

STATUSES = [
    "Принято в работу",
    "Запрос цены",
    "Отправлено предложение",
    "На согласовании у Заказчика",
    "Отправлено",
]
CUSTOMERS = [f"Заказчик {i}" for i in range(1, 11)]
SUPPLIERS = [f"Поставщик {i}" for i in range(1, 4)]

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BASE_FONT = Font(name=FONT_NAME)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_COLORS = {
    "Принято в работу": "BDD7EE",          # светло-синий
    "Запрос цены": "FFE699",               # светло-жёлтый
    "Отправлено предложение": "D9C6EC",    # светло-фиолетовый
    "На согласовании у Заказчика": "C6E0B4",  # светло-зелёный
    "Отправлено": "D9D9D9",                # серый (гаснет)
}
OVERDUE_FILL = "F8696B"
MISSING_DATE_FILL = "FFEB9C"

wb = Workbook()

# ---------------------------------------------------------------------------
# Лист "Справочники"
# ---------------------------------------------------------------------------
ref = wb.active
ref.title = "Справочники"

ref["B1"] = "Заказчики"
ref["C1"] = "Поставщики"
ref["D1"] = "Статусы"
ref["E1"] = "Порог просрочки, дней"
ref["F1"] = 5

for cell in ("B1", "C1", "D1", "E1"):
    ref[cell].font = HEADER_FONT
    ref[cell].fill = HEADER_FILL
    ref[cell].alignment = Alignment(horizontal="center", wrap_text=True)

for i, name in enumerate(CUSTOMERS, start=2):
    ref.cell(row=i, column=2, value=name)
for i, name in enumerate(SUPPLIERS, start=2):
    ref.cell(row=i, column=3, value=name)
for i, name in enumerate(STATUSES, start=2):
    ref.cell(row=i, column=4, value=name)

ref["F1"].font = Font(name=FONT_NAME, bold=True, color="C00000")
ref["F1"].alignment = Alignment(horizontal="center")

for row in ref.iter_rows(min_row=1, max_row=11, min_col=2, max_col=6):
    for cell in row:
        if cell.font.name is None or cell.font != HEADER_FONT:
            cell.font = BASE_FONT

ref.column_dimensions["A"].width = 3
for col in "BCDEF":
    ref.column_dimensions[col].width = 24

# Именованные диапазоны
def add_defined_name(name, ref_formula):
    wb.defined_names[name] = DefinedName(name, attr_text=ref_formula)

add_defined_name("Заказчики", f"'Справочники'!$B$2:$B${1+len(CUSTOMERS)}")
add_defined_name("Поставщики", f"'Справочники'!$C$2:$C${1+len(SUPPLIERS)}")
add_defined_name("Статусы", f"'Справочники'!$D$2:$D${1+len(STATUSES)}")
add_defined_name("Порог_просрочки", "'Справочники'!$F$1")

# ---------------------------------------------------------------------------
# Лист "Заявки"
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Заявки")

headers = ["№", "Дата поступления", "Заказчик", "Поставщик", "Текущий статус"] + STATUSES + [
    "Дней на текущем статусе", "Комментарий"
]
for col, title in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=title)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

# Индексы колонок
COL_NUM = 1
COL_DATE_IN = 2
COL_CUSTOMER = 3
COL_SUPPLIER = 4
COL_STATUS = 5
COL_STATUS_DATES_START = 6                      # F
COL_STATUS_DATES_END = COL_STATUS_DATES_START + len(STATUSES) - 1  # J
COL_DAYS = COL_STATUS_DATES_END + 1             # K
COL_COMMENT = COL_DAYS + 1                      # L

STATUS_DATE_LETTERS = [get_column_letter(c) for c in range(COL_STATUS_DATES_START, COL_STATUS_DATES_END + 1)]
STATUS_HEADER_RANGE = f"${STATUS_DATE_LETTERS[0]}$1:${STATUS_DATE_LETTERS[-1]}$1"

date_style = "DD.MM.YYYY"

for r in range(FIRST_ROW, LAST_ROW + 1):
    num_cell = ws.cell(row=r, column=COL_NUM)
    num_cell.value = f'=IF({get_column_letter(COL_DATE_IN)}{r}="","",ROW()-{FIRST_ROW - 1})'
    num_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=r, column=COL_DATE_IN).number_format = date_style
    for c in range(COL_STATUS_DATES_START, COL_STATUS_DATES_END + 1):
        ws.cell(row=r, column=c).number_format = date_style

    status_col_letter = get_column_letter(COL_STATUS)
    days_cell = ws.cell(row=r, column=COL_DAYS)
    row_range_start = get_column_letter(COL_STATUS_DATES_START)
    row_range_end = get_column_letter(COL_STATUS_DATES_END)
    days_cell.value = (
        f'=IF({status_col_letter}{r}="","",'
        f'IF(COUNTIF({STATUS_HEADER_RANGE},{status_col_letter}{r})=0,"",'
        f'IF(INDEX({row_range_start}{r}:{row_range_end}{r},MATCH({status_col_letter}{r},{STATUS_HEADER_RANGE},0))="","",'
        f'TODAY()-INDEX({row_range_start}{r}:{row_range_end}{r},MATCH({status_col_letter}{r},{STATUS_HEADER_RANGE},0)))))'
    )
    days_cell.alignment = Alignment(horizontal="center")

    for c in range(1, COL_COMMENT + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = BASE_FONT
        cell.border = BORDER

# Ширины колонок
widths = {
    "A": 6, "B": 14, "C": 18, "D": 16, "E": 22,
    "F": 14, "G": 14, "H": 16, "I": 18, "J": 14,
    "K": 14, "L": 30,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Выпадающие списки
dv_customer = DataValidation(type="list", formula1="=Заказчики", allow_blank=True, showDropDown=False)
dv_supplier = DataValidation(type="list", formula1="=Поставщики", allow_blank=True, showDropDown=False)
dv_status = DataValidation(type="list", formula1="=Статусы", allow_blank=True, showDropDown=False)
for dv in (dv_customer, dv_supplier, dv_status):
    dv.error = "Выберите значение из списка"
    dv.errorTitle = "Некорректное значение"
    ws.add_data_validation(dv)

dv_customer.add(f"C{FIRST_ROW}:C{LAST_ROW}")
dv_supplier.add(f"D{FIRST_ROW}:D{LAST_ROW}")
dv_status.add(f"E{FIRST_ROW}:E{LAST_ROW}")

# Закреплённая шапка + автофильтр
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{LAST_ROW}"

full_range = f"A{FIRST_ROW}:L{LAST_ROW}"
status_col_range = f"$E{FIRST_ROW}:$E{LAST_ROW}"

# 1) Статус "Отправлено" — гашение строки (наивысший приоритет)
closed_fill = PatternFill("solid", fgColor=STATUS_COLORS["Отправлено"])
closed_font = Font(name=FONT_NAME, color="808080", italic=True)
rule_closed = FormulaRule(
    formula=[f'$E{FIRST_ROW}="Отправлено"'],
    fill=closed_fill,
    font=closed_font,
)
ws.conditional_formatting.add(full_range, rule_closed)

# 2) Цвет по остальным статусам
for status in STATUSES:
    if status == "Отправлено":
        continue
    fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
    rule = FormulaRule(formula=[f'$E{FIRST_ROW}="{status}"'], fill=fill)
    ws.conditional_formatting.add(full_range, rule)

# 3) Дата статуса не проставлена, а статус выбран — жёлтая подсветка ячейки даты
yellow_fill = PatternFill("solid", fgColor=MISSING_DATE_FILL)
for letter in STATUS_DATE_LETTERS:
    col_range = f"{letter}{FIRST_ROW}:{letter}{LAST_ROW}"
    rule = FormulaRule(
        formula=[f'AND($E{FIRST_ROW}={letter}$1,{letter}{FIRST_ROW}="")'],
        fill=yellow_fill,
    )
    ws.conditional_formatting.add(col_range, rule)

# 4) Просрочка по дням на статусе — красная подсветка колонки K
red_fill = PatternFill("solid", fgColor=OVERDUE_FILL)
days_range = f"K{FIRST_ROW}:K{LAST_ROW}"
rule_overdue = FormulaRule(
    formula=[f'AND(K{FIRST_ROW}<>"",K{FIRST_ROW}>Порог_просрочки)'],
    fill=red_fill,
    font=Font(name=FONT_NAME, bold=True, color="9C0006"),
)
ws.conditional_formatting.add(days_range, rule_overdue)

ws.sheet_view.showGridLines = False

# ---------------------------------------------------------------------------
# Лист "Сводка"
# ---------------------------------------------------------------------------
sm = wb.create_sheet("Сводка")

sm["A1"] = "Сводка по заявкам"
sm["A1"].font = Font(name=FONT_NAME, bold=True, size=14)

sm["A3"] = "Статус"
sm["B3"] = "Количество"
for cell in ("A3", "B3"):
    sm[cell].font = HEADER_FONT
    sm[cell].fill = HEADER_FILL
    sm[cell].alignment = Alignment(horizontal="center")

status_range = f"Заявки!$E${FIRST_ROW}:$E${LAST_ROW}"
for i, status in enumerate(STATUSES, start=4):
    sm.cell(row=i, column=1, value=status).font = BASE_FONT
    cnt_cell = sm.cell(row=i, column=2)
    cnt_cell.value = f'=COUNTIF({status_range},A{i})'
    cnt_cell.font = BASE_FONT
    cnt_cell.alignment = Alignment(horizontal="center")

last_status_row = 3 + len(STATUSES)

sm.cell(row=last_status_row + 2, column=1, value="Общее количество заявок").font = Font(name=FONT_NAME, bold=True)
total_cell = sm.cell(row=last_status_row + 2, column=2)
total_cell.value = f'=COUNTA({status_range})'
total_cell.font = Font(name=FONT_NAME, bold=True)
total_cell.alignment = Alignment(horizontal="center")

sm.cell(row=last_status_row + 3, column=1, value="Количество просроченных заявок").font = Font(name=FONT_NAME, bold=True, color="C00000")
overdue_cell = sm.cell(row=last_status_row + 3, column=2)
days_full_range = f"Заявки!$K${FIRST_ROW}:$K${LAST_ROW}"
overdue_cell.value = f'=COUNTIF({days_full_range},">"&Порог_просрочки)'
overdue_cell.font = Font(name=FONT_NAME, bold=True, color="C00000")
overdue_cell.alignment = Alignment(horizontal="center")

sm.column_dimensions["A"].width = 32
sm.column_dimensions["B"].width = 16

# Диаграмма распределения заявок по статусам
chart = BarChart()
chart.title = "Распределение заявок по статусам"
chart.y_axis.title = "Количество"
chart.x_axis.title = "Статус"
chart.style = 10
chart.width = 18
chart.height = 10

data = Reference(sm, min_col=2, min_row=3, max_row=last_status_row)
cats = Reference(sm, min_col=1, min_row=4, max_row=last_status_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None

sm.add_chart(chart, "D3")

# Порядок листов: Заявки первым видимым рабочим листом
wb.move_sheet("Заявки", offset=-1)

wb.save("Трекер_заявок.xlsx")
print("OK: Трекер_заявок.xlsx создан")
